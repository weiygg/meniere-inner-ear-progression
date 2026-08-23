from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook


STEPS = [
    "00_config.py",
    "01_inventory_files.py",
    "02_audit_clinical_data.py",
    "03_link_patient_ear.py",
    "04_segmentation_qc.py",
    "05_extract_inner_ear_morphometry.py",
    "11_run_pebm.py",
]


def _sheet_rows(path: Path, sheet: str):
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = workbook[sheet].iter_rows(values_only=True)
    headers = next(rows)
    return [dict(zip(headers, row)) for row in rows]


def write_first_round_summary(config_path: Path) -> str:
    config = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    output = Path(config["OUTPUT_ROOT"])
    clinical = {row["metric"]: row["value"] for row in _sheet_rows(output / "01_data_audit" / "clinical_audit_summary.xlsx", "summary")}
    variables = _sheet_rows(output / "01_data_audit" / "variable_dictionary.xlsx", "variables")
    missing = _sheet_rows(output / "01_data_audit" / "missingness_summary.xlsx", "ear_records")
    qc = {row["metric"]: row["value"] for row in _sheet_rows(output / "03_morphometry" / "qc_metrics.xlsx", "summary")}
    morph = _sheet_rows(output / "03_morphometry" / "morphometry_features.xlsx", "features")
    centerlines = _sheet_rows(output / "03_morphometry" / "morphometry_features.xlsx", "canal_centerlines")
    reproduction = _sheet_rows(output / "04_pebm" / "reproduction_test_results.xlsx", "sequence_tests")
    def unique(values):
        return list(dict.fromkeys(values))

    primary = unique([str(row["original_column"]) for row in variables if row.get("included_in_primary_pebm")])
    endotype = unique([str(row["original_column"]) for row in variables if row.get("included_in_endotype")])
    excluded = unique([str(row["original_column"]) for row in variables if row.get("variable_role") in {"A_static_anatomical_endotype", "C_fluctuating_symptom", "D_independent_validation"}])
    missing_ranked = sorted(missing, key=lambda row: float(row["missing_rate"] or 0), reverse=True)
    confirmation = (output / "NEED_CONFIRMATION.md").read_text(encoding="utf-8")
    blocking_n = sum(line.startswith("## ") for line in confirmation.splitlines())
    truths = all(bool(row["truth_recovered"]) for row in reproduction)
    centerline_pass = sum(row.get("centerline_status") == "pass" for row in centerlines)
    lines = [
        "# 首轮执行摘要", "",
        f"1. 患者/耳侧：共 {int(clinical['unique_patient'])} 名唯一患者、{int(clinical['baseline_ear_rows'])} 条基线耳侧记录；{int(clinical['patients_with_followup'])} 名患者存在重复访视。",
        f"2. 患者级主队列代理：{int(clinical.get('baseline_affected_ear_proxy', 0))} 条患耳代理记录；{int(clinical['primary_patient_level_n'])} 名患者可唯一确定索引耳代理，另有 {int(clinical['index_ear_unresolved_n'])} 名尚未解决。",
        f"3. 动态客观 P-EBM 候选变量（编码确认前仅列候选）：{', '.join(primary) if primary else '无'}。",
        f"4. 临床表中当前识别的静态解剖内表型变量：{', '.join(endotype) if endotype else '无'}；分割派生几何特征另表存储。",
        f"5. 不适合纳入主要客观 P-EBM 的变量包括静态解剖、波动性症状及预留的独立验证变量：{', '.join(excluded) if excluded else '无'}。",
        f"6. 缺失情况：最高观察缺失率为 {missing_ranked[0]['variable']}（{float(missing_ranked[0]['missing_rate']):.1%}）；全部变量见 missingness_summary.xlsx。",
        f"7. 分割 QC：{int(qc['pass'])}/{int(qc['mask_files'])} 个掩膜通过（{float(qc['pass_rate']):.1%}）；{int(qc['flagged'])} 个被标记且未静默排除。",
        f"8. 官方 P-EBM 复现：固定上游提交上的串行事件与同时事件模拟均{'成功' if truths else '失败'}。",
        f"9. 形态学/试运行：{len(morph)} 个掩膜已获得基础特征；{centerline_pass}/{len(centerlines)} 个半规管掩膜成功获得中心线。因关键定义仍未确认，未运行真实数据 P-EBM 试点。",
        f"10. 下一步确认：NEED_CONFIRMATION.md 列出 {blocking_n} 个阻断主题，重点为分割-临床 ID 映射、权威分割批次、结构命名、编码/方向、索引耳及事件参考组。",
    ]
    text = "\n".join(lines) + "\n"
    (output / "first_round_summary.md").write_text(text, encoding="utf-8")
    return text


def main() -> int:
    parser = argparse.ArgumentParser(description="Run first-round MD progression workflow")
    parser.add_argument("--config", type=Path, default=Path("config/project_config.yaml"))
    parser.add_argument("--skip-montage", action="store_true")
    args = parser.parse_args()
    root = Path(__file__).resolve().parent
    for script in STEPS:
        cmd = [sys.executable, str(root / "src" / script), "--config", str(args.config.resolve())]
        if script == "04_segmentation_qc.py" and args.skip_montage:
            cmd.append("--skip-montage")
        print(f"[run_all] {' '.join(cmd)}", flush=True)
        completed = subprocess.run(cmd, cwd=root)
        if completed.returncode:
            print(f"[run_all] stopped at {script} (exit={completed.returncode})", file=sys.stderr)
            return completed.returncode
    print(write_first_round_summary(args.config.resolve()), flush=True)
    print("[run_all] first-round workflow completed")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
