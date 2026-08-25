"""Create PHI-safe repository and source-data audit manifests.

The script records file metadata, hashes, workbook schemas, and aggregate counts.
It never exports clinical rows or DICOM metadata. Raw inputs are read-only.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import subprocess
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


AUDIT_EXTENSIONS = {
    ".csv",
    ".tsv",
    ".xls",
    ".xlsx",
    ".xlsm",
    ".doc",
    ".docx",
    ".pdf",
    ".ppt",
    ".pptx",
    ".ipynb",
    ".r",
    ".rmd",
    ".py",
    ".ps1",
    ".md",
    ".json",
    ".yaml",
    ".yml",
}

SKIP_PARTS = {".git", ".venv", ".agents", ".codex", "tmp", "__pycache__", ".pytest_cache"}
PROTECTED_TOP_LEVEL = {
    "data",
    "analysis_out",
    "results_md_progression",
    "seg3",
    "seg4",
    "xjj内耳分割",
    "xjj内耳分割2",
    "archive",
}
DIRECT_IDENTIFIER_HEADERS = {
    "姓名",
    "患者姓名",
    "电话",
    "病案号",
    "出生日期",
    "扫描日期",
}
NUMERIC_AUDIT_COLUMNS = {
    "0.5kHZ",
    "1kHZ",
    "2kHZ",
    "3kHZ",
    "4kHZ",
    "PTA",
    "CochEH",
    "VestEH",
    "VA",
    "ES/ED",
    "stage（AAO-HNS）",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def relative(root: Path, path: Path) -> str:
    return path.relative_to(root).as_posix()


def git(root: Path, *args: str) -> str:
    result = subprocess.run(
        ["git", *args],
        cwd=root,
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return result.stdout.strip()


def write_csv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def classify(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls", ".xlsm", ".csv", ".tsv"}:
        return "tabular_data_or_output"
    if suffix in {".doc", ".docx", ".pdf", ".ppt", ".pptx"}:
        return "document_or_figure"
    if suffix in {".py", ".r", ".rmd", ".ipynb", ".ps1"}:
        return "analysis_code"
    if suffix in {".yaml", ".yml", ".json"}:
        return "configuration_or_machine_readable_output"
    return "documentation_or_other"


def iter_audit_files(root: Path) -> Iterable[Path]:
    for path in root.rglob("*"):
        if not path.is_file() or any(part in SKIP_PARTS for part in path.relative_to(root).parts):
            continue
        if path.relative_to(root).parts[:2] == ("data", "manifests"):
            continue
        if path.suffix.lower() not in AUDIT_EXTENSIONS:
            continue
        # Avoid recording paths nested under subject/study folders. Their contents are
        # represented only by aggregate top-level storage summaries.
        parts = path.relative_to(root).parts[:-1]
        if any(re.fullmatch(r"(?:sub\d+|\d{3}(?:-\d+[a-z])?)", part, re.I) for part in parts):
            continue
        if any(re.match(r"MR\d+", part, re.I) for part in parts):
            continue
        yield path


def tracked_inventory(root: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in git(root, "ls-files").splitlines():
        if not item:
            continue
        path = root / item
        stat = path.stat()
        rows.append(
            {
                "path": item,
                "bytes": stat.st_size,
                "sha256": sha256(path),
                "modified_time": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                "category": classify(path),
                "git_tracked": "yes",
            }
        )
    return rows


def local_storage_summary(root: Path) -> list[dict[str, Any]]:
    summary: dict[str, dict[str, int]] = defaultdict(lambda: {"files": 0, "bytes": 0})
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        parts = path.relative_to(root).parts
        if not parts or parts[0] in SKIP_PARTS:
            continue
        top = parts[0]
        summary[top]["files"] += 1
        summary[top]["bytes"] += path.stat().st_size
    return [
        {
            "top_level": top,
            "file_count": values["files"],
            "bytes": values["bytes"],
            "protected_or_generated": "yes" if top in PROTECTED_TOP_LEVEL else "no",
            "inventory_scope": "aggregate_only" if top in PROTECTED_TOP_LEVEL else "file_level_for_git_content",
        }
        for top, values in sorted(summary.items(), key=lambda item: item[1]["bytes"], reverse=True)
    ]


def normalize_duplicate_key(path: Path) -> str:
    stem = path.stem.lower()
    stem = re.sub(r"(?:19|20)\d{6,12}", "<date>", stem)
    stem = re.sub(r"(?:^|[_-])v?\d+(?:\.\d+)*(?=$|[_-])", "_", stem)
    stem = re.sub(r"(?:final|updated|current|legacy|old|copy|checkpoint)", "", stem)
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", stem)


def audit_artifacts(root: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    by_hash: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    tracked_paths = set(git(root, "ls-files").splitlines())
    for path in sorted(iter_audit_files(root)):
        stat = path.stat()
        digest = sha256(path)
        row = {
            "path": relative(root, path),
            "bytes": stat.st_size,
            "sha256": digest,
            "created_time_local_fs": datetime.fromtimestamp(stat.st_ctime).isoformat(timespec="seconds"),
            "modified_time": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            "category": classify(path),
            "git_tracked": "yes" if relative(root, path) in tracked_paths else "no",
        }
        rows.append(row)
        by_hash[digest].append(row)
        key = normalize_duplicate_key(path)
        if key:
            by_name[key].append(row)

    exact: list[dict[str, Any]] = []
    group_id = 0
    for digest, members in sorted(by_hash.items()):
        if len(members) < 2:
            continue
        group_id += 1
        for member in members:
            exact.append(
                {
                    "duplicate_group": group_id,
                    "sha256": digest,
                    "path": member["path"],
                    "bytes": member["bytes"],
                }
            )

    near: list[dict[str, Any]] = []
    near_group = 0
    for key, members in sorted(by_name.items()):
        hashes = {member["sha256"] for member in members}
        if len(members) < 2 or len(hashes) < 2:
            continue
        near_group += 1
        for member in members:
            near.append(
                {
                    "near_duplicate_group": near_group,
                    "normalized_name": key,
                    "path": member["path"],
                    "bytes": member["bytes"],
                    "sha256": member["sha256"],
                }
            )
    return rows, exact, near


def as_number(value: Any) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def workbook_audit(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    workbook_values = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    for worksheet, values_sheet in zip(workbook.worksheets, workbook_values.worksheets):
        rows = list(worksheet.iter_rows(values_only=True))
        value_rows = list(values_sheet.iter_rows(values_only=True))
        headers = [str(value).strip() if value is not None else "" for value in (rows[0] if rows else [])]
        header_index = {header: index for index, header in enumerate(headers) if header}
        data_rows = rows[1:]
        aggregate: dict[str, Any] = {}
        for column in sorted(NUMERIC_AUDIT_COLUMNS & set(header_index)):
            values = [as_number(row[header_index[column]]) if header_index[column] < len(row) else None for row in data_rows]
            numeric = [value for value in values if value is not None]
            aggregate[column] = {
                "nonmissing_n": len(numeric),
                "min": min(numeric) if numeric else None,
                "max": max(numeric) if numeric else None,
            }

        id_summary: dict[str, Any] = {}
        if "ID" in header_index:
            ids = [row[header_index["ID"]] for row in data_rows if header_index["ID"] < len(row) and row[header_index["ID"]] not in (None, "")]
            id_summary = {"nonmissing_n": len(ids), "unique_n": len({str(value) for value in ids})}

        side_summary: dict[str, int] = {}
        if "side" in header_index:
            side_summary = dict(
                Counter(
                    str(row[header_index["side"]]).strip()
                    for row in data_rows
                    if header_index["side"] < len(row) and row[header_index["side"]] not in (None, "")
                )
            )

        pta_check: dict[str, Any] = {
            "four_frequency_complete_rows": 0,
            "pta_formula_rows": 0,
            "pta_formula_matches_0.5_1_2_3khz_columns": 0,
            "stored_numeric_pta_rows": 0,
            "stored_numeric_pta_matches_within_1e-6": 0,
            "max_absolute_difference_for_stored_numeric_pta": None,
        }
        required = ["0.5kHZ", "1kHZ", "2kHZ", "3kHZ", "PTA"]
        differences: list[float] = []
        if all(column in header_index for column in required):
            complete_frequency_rows = 0
            formula_rows = 0
            expected_formula_rows = 0
            start_column = get_column_letter(header_index["0.5kHZ"] + 1)
            end_column = get_column_letter(header_index["3kHZ"] + 1)
            for excel_row, row in enumerate(data_rows, start=2):
                frequencies = [
                    as_number(row[header_index[column]]) if header_index[column] < len(row) else None
                    for column in required[:4]
                ]
                if any(value is None for value in frequencies):
                    continue
                complete_frequency_rows += 1
                recomputed = sum(frequencies) / 4
                pta_cell = row[header_index["PTA"]] if header_index["PTA"] < len(row) else None
                if isinstance(pta_cell, str) and pta_cell.startswith("="):
                    formula_rows += 1
                    normalized = pta_cell.replace("$", "").replace(" ", "").upper()
                    expected = f"=AVERAGE({start_column}{excel_row}:{end_column}{excel_row})"
                    if normalized == expected:
                        expected_formula_rows += 1
                else:
                    stored = as_number(pta_cell)
                    if stored is not None:
                        differences.append(abs(recomputed - stored))
            pta_check = {
                "four_frequency_complete_rows": complete_frequency_rows,
                "pta_formula_rows": formula_rows,
                "pta_formula_matches_0.5_1_2_3khz_columns": expected_formula_rows,
                "stored_numeric_pta_rows": len(differences),
                "stored_numeric_pta_matches_within_1e-6": sum(value <= 1e-6 for value in differences),
                "max_absolute_difference_for_stored_numeric_pta": max(differences) if differences else None,
            }

        date_ranges: dict[str, dict[str, str | None]] = {}
        for column in ("起病时间", "扫描日期"):
            if column not in header_index:
                continue
            dates = [
                parsed
                for row in data_rows
                if header_index[column] < len(row)
                for parsed in [iso_date(row[header_index[column]])]
                if parsed is not None
            ]
            date_ranges[column] = {"min": min(dates) if dates else None, "max": max(dates) if dates else None, "nonmissing_n": len(dates)}

        formula_count = sum(
            isinstance(cell, str) and cell.startswith("=")
            for row in rows
            for cell in row
        )
        formula_error_count = sum(
            isinstance(cell, str) and cell.startswith("#")
            for row in value_rows
            for cell in row
        )
        sheets.append(
            {
                "sheet_index": len(sheets) + 1,
                "sheet_name": worksheet.title,
                "data_rows": max(len(rows) - 1, 0),
                "columns": len(headers),
                "headers": headers,
                "direct_identifier_headers": sorted(DIRECT_IDENTIFIER_HEADERS & set(headers)),
                "id_summary": id_summary,
                "side_value_counts": side_summary,
                "numeric_column_summary": aggregate,
                "pta_0.5_1_2_3khz_audit": pta_check,
                "has_4khz_column": "4kHZ" in header_index,
                "date_ranges": date_ranges,
                "formula_count": formula_count,
                "cached_formula_error_count": formula_error_count,
            }
        )
    return {
        "source_name": path.name,
        "bytes": path.stat().st_size,
        "sha256": sha256(path),
        "phi_status": "direct_identifiers_present",
        "included_in_git": False,
        "sheets": sheets,
    }


def raw_manifest(root: Path, clinical_path: Path, clinical_audit: dict[str, Any]) -> list[dict[str, Any]]:
    sheet_summary = "; ".join(
        f"{sheet['sheet_name']}:{sheet['data_rows']}x{sheet['columns']}"
        for sheet in clinical_audit["sheets"]
    )
    date_ranges = []
    for sheet in clinical_audit["sheets"]:
        for label, values in sheet["date_ranges"].items():
            if values["min"] or values["max"]:
                date_ranges.append(f"{sheet['sheet_name']}.{label}:{values['min']}..{values['max']}")
    rows = [
        {
            "source_name": clinical_path.name,
            "logical_source": "protected local clinical workbook",
            "sha256": clinical_audit["sha256"],
            "rows": sheet_summary,
            "columns": "13/37/23 by sheet",
            "date_range": "; ".join(date_ranges) if date_ranges else "not available",
            "identifier_structure": "site-specific numeric ID; separate patient-level number; direct identifiers also present",
            "analysis_level": "ear-level sheets 1 and 3; patient-level sheet 2; repeat visits present",
            "phi_status": "PHI/direct identifiers - restricted",
            "included_in_git": "no",
        }
    ]
    archive_roles = {
        "丽水-xjj内耳分割4.rar": ("primary Lishui development segmentation archive", "patient/ear/structure"),
        "浙二1-1.rar": ("external validation 1 imaging archive batch", "patient/study/series"),
        "浙二1-2.rar": ("external validation 1 imaging archive batch", "patient/study/series"),
        "浙二2-1.rar": ("external validation 2 imaging archive batch", "patient/study/series"),
        "浙二2-2.rar": ("external validation 2 imaging archive batch", "patient/study/series"),
        "浙二2例新.rar": ("external validation 2 imaging archive addendum", "patient/study/series"),
        "中心2外部验证1.rar": ("external validation 1 manual-mask archive", "center/patient/ear/structure"),
        "中心3外部验证2.rar": ("external validation 2 manual-mask archive", "center/patient/ear/structure"),
    }
    for path in sorted((root / "data").rglob("*.rar")):
        role, level = archive_roles.get(path.name, ("protected imaging archive", "unknown - requires adjudication"))
        rows.append(
            {
                "source_name": path.name,
                "logical_source": role,
                "sha256": sha256(path),
                "rows": "not applicable",
                "columns": "not applicable",
                "date_range": "not extracted at audit checkpoint",
                "identifier_structure": "archive names/folders may contain study or patient identifiers",
                "analysis_level": level,
                "phi_status": "potentially identifiable imaging - restricted",
                "included_in_git": "no",
            }
        )
    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument(
        "--clinical-workbook",
        type=Path,
        default=Path("data/clinical/MD患者评估20260713.xlsx"),
    )
    parser.add_argument("--output-dir", type=Path, default=Path("data/manifests"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = args.root.resolve()
    clinical_path = args.clinical_workbook
    if not clinical_path.is_absolute():
        clinical_path = root / clinical_path
    output_dir = args.output_dir
    if not output_dir.is_absolute():
        output_dir = root / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    tracked = tracked_inventory(root)
    storage = local_storage_summary(root)
    artifacts, duplicates, near_duplicates = audit_artifacts(root)
    clinical = workbook_audit(clinical_path)
    sources = raw_manifest(root, clinical_path, clinical)

    write_csv(output_dir / "repository_file_inventory.csv", list(tracked[0]), tracked)
    write_csv(output_dir / "local_storage_summary.csv", list(storage[0]), storage)
    write_csv(output_dir / "legacy_analysis_hashes.csv", list(artifacts[0]), artifacts)
    write_csv(
        output_dir / "duplicate_file_groups.csv",
        ["duplicate_group", "sha256", "path", "bytes"],
        duplicates,
    )
    write_csv(
        output_dir / "near_duplicate_candidates.csv",
        ["near_duplicate_group", "normalized_name", "path", "bytes", "sha256"],
        near_duplicates,
    )
    write_csv(output_dir / "raw_data_manifest.csv", list(sources[0]), sources)
    (output_dir / "clinical_workbook_audit.json").write_text(
        json.dumps(clinical, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    context = {
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "branch": git(root, "branch", "--show-current"),
        "commit": git(root, "rev-parse", "HEAD"),
        "git_status_porcelain": git(root, "status", "--porcelain"),
        "tracked_file_count": len(tracked),
        "audited_artifact_count": len(artifacts),
        "exact_duplicate_group_count": len({row["duplicate_group"] for row in duplicates}),
        "near_duplicate_group_count": len({row["near_duplicate_group"] for row in near_duplicates}),
        "safety": "No patient rows or DICOM metadata exported.",
    }
    (output_dir / "audit_context.json").write_text(
        json.dumps(context, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(context, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
