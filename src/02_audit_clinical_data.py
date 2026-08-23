from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from mdp_utils import (
    load_config,
    provisional_variable_role,
    read_ear_records,
    setup_logger,
    write_csv,
    write_xlsx,
)


def infer_type_and_values(values):
    present = [value for value in values if value not in (None, "")]
    if not present:
        return "unknown", ""
    if all(isinstance(value, bool) for value in present):
        data_type = "boolean"
    elif all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in present):
        data_type = "integer" if all(float(value).is_integer() for value in present) else "continuous"
    elif all(isinstance(value, (date, datetime)) for value in present):
        data_type = "date"
    else:
        data_type = "categorical" if len({str(value) for value in present}) <= 20 else "text"
    unique = sorted({str(value) for value in present})
    allowed = "; ".join(unique) if len(unique) <= 20 else ""
    return data_type, allowed


def english_description(name: str) -> str:
    mapping = {
        "age": "Age",
        "sex": "Sex",
        "side": "Ear side",
        "PTA": "Pure-tone average",
        "CochEH": "Cochlear endolymphatic hydrops grade",
        "VestEH": "Vestibular endolymphatic hydrops grade",
        "VA": "Vestibular aqueduct imaging grade",
        "ES/ED": "Endolymphatic sac/duct imaging grade",
        "THI": "Tinnitus Handicap Inventory",
        "VADL": "Vestibular Activities of Daily Living score",
        "耳闷": "Aural fullness",
        "stage（AAO-HNS）": "AAO-HNS hearing stage",
    }
    return mapping.get(name, name)


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit clinical data and build first-round cohorts")
    parser.add_argument("--config", type=Path, required=True)
    args = parser.parse_args()
    _, paths = load_config(args.config)
    log = setup_logger("clinical", paths.logs / "02_audit_clinical_data.log")
    records = read_ear_records(paths.clinical_table)
    output = paths.output_root / "01_data_audit"
    output.mkdir(parents=True, exist_ok=True)

    fields = sorted({key for record in records for key in record})
    missing = []
    for field in fields:
        valid_n = sum(record.get(field) not in (None, "") for record in records)
        missing.append([field, valid_n, len(records) - valid_n, (len(records) - valid_n) / len(records)])
    workbook = load_workbook(paths.clinical_table, read_only=True, data_only=False)
    source_missing = []
    for worksheet in workbook.worksheets:
        source_rows = list(worksheet.iter_rows(values_only=True))
        if not source_rows:
            continue
        headers = source_rows[0]
        data_rows = source_rows[1:]
        for column_index, header in enumerate(headers):
            if header is None:
                continue
            values = [row[column_index] if column_index < len(row) else None for row in data_rows]
            valid_n = sum(value not in (None, "") for value in values)
            source_missing.append([worksheet.title, str(header).strip(), valid_n, len(values) - valid_n, (len(values) - valid_n) / len(values) if values else None])
    write_xlsx(output / "missingness_summary.xlsx", {
        "ear_records": (["variable", "valid_n", "missing_n", "missing_rate"], missing),
        "source_columns": (["source_sheet", "original_column", "valid_n", "missing_n", "missing_rate"], source_missing),
    })

    duplicates = defaultdict(list)
    for record in records:
        duplicates[(record["visit_id"], record["ear_side"])].append(record["source_row"])
    conflicts = [[key[0], key[1], ";".join(map(str, rows)), "duplicate visit-ear"] for key, rows in duplicates.items() if len(rows) > 1]
    write_xlsx(output / "duplicate_and_conflict_report.xlsx", {"duplicates": (["visit_id", "ear_side", "source_rows", "issue"], conflicts)})

    dictionary = []
    seen = set()
    for worksheet in workbook.worksheets:
        worksheet_rows = list(worksheet.iter_rows(values_only=True))
        if not worksheet_rows:
            continue
        for column_index, cell_value in enumerate(worksheet_rows[0]):
            if cell_value is None:
                continue
            original = str(cell_value).strip()
            key = (worksheet.title, original)
            if key in seen:
                continue
            seen.add(key)
            role, primary, extended, endotype, note = provisional_variable_role(original)
            source_values = [row[column_index] if column_index < len(row) else None for row in worksheet_rows[1:]]
            data_type, allowed = infer_type_and_values(source_values)
            present_n = sum(value not in (None, "") for value in source_values)
            missing_rate = (len(source_values) - present_n) / len(source_values) if source_values else None
            direction = "requires confirmation" if role in {"B_dynamic_objective_biomarker", "D_independent_validation"} else "not applicable"
            unit = "requires confirmation" if original in {"PTA", "0.5kHZ", "1kHZ", "2kHZ", "3kHZ"} else ""
            dictionary.append([
                original, original, original, english_description(original), data_type, unit, allowed,
                direction, role, primary, extended, endotype, role == "D_independent_validation",
                missing_rate, note, worksheet.title,
            ])
    dictionary_headers = [
        "original_column", "standard_name", "Chinese_description", "English_description",
        "data_type", "unit", "allowed_values", "direction_of_abnormality", "variable_role",
        "included_in_primary_pebm", "included_in_extended_pebm", "included_in_endotype",
        "used_as_validation_outcome", "missing_rate", "notes", "source_sheet",
    ]
    write_xlsx(output / "variable_dictionary.xlsx", {"variables": (dictionary_headers, dictionary)})

    patients = {record["patient_id"] for record in records}
    ears = {record["ear_id"] for record in records}
    followup_patients = {record["patient_id"] for record in records if record["is_followup"]}
    baseline = [record for record in records if not record["is_followup"]]
    affected = sum(bool(record["affected_ear_proxy"]) for record in baseline)
    site_counts = Counter(record["source_site"] for record in baseline)
    summary = [
        ["unique_patient", len(patients)], ["unique_ear", len(ears)],
        ["baseline_ear_rows", len(baseline)], ["baseline_affected_ear_proxy", affected],
        ["patients_with_followup", len(followup_patients)],
        *[[f"baseline_ear_rows_{site}", count] for site, count in site_counts.items()],
    ]
    write_xlsx(output / "clinical_audit_summary.xlsx", {
        "summary": (["metric", "value"], summary),
        "normalized_ear_records": (fields, ([record.get(field) for field in fields] for record in records)),
    })

    cohort = paths.output_root / "02_cohort"
    cohort.mkdir(parents=True, exist_ok=True)
    write_csv(cohort / "all_ear_sensitivity_cohort.csv", fields, ([record.get(field) for field in fields] for record in baseline))
    by_patient = defaultdict(list)
    for record in baseline:
        by_patient[record["patient_id"]].append(record)
    primary, unresolved = [], []
    for patient_id, patient_rows in by_patient.items():
        candidates = [record for record in patient_rows if record["affected_ear_proxy"]]
        if len(candidates) == 1:
            primary.append(candidates[0])
        else:
            unresolved.append([patient_id, len(patient_rows), len(candidates), "affected/index ear not uniquely encoded"])
    summary.extend([["primary_patient_level_n", len(primary)], ["index_ear_unresolved_n", len(unresolved)]])
    write_xlsx(output / "clinical_audit_summary.xlsx", {
        "summary": (["metric", "value"], summary),
        "normalized_ear_records": (fields, ([record.get(field) for field in fields] for record in records)),
    })
    write_csv(cohort / "primary_patient_level_cohort.csv", fields, ([record.get(field) for field in fields] for record in primary))
    write_csv(cohort / "unilateral_md_affected_ear.csv", fields, [])
    write_csv(cohort / "unilateral_md_contralateral_ear.csv", fields, [])
    blocked = [
        ["unilateral_md_affected_ear.csv", "not constructed", "unilateral/bilateral status and affected side are not uniquely encoded"],
        ["unilateral_md_contralateral_ear.csv", "not constructed", "contralateral ear cannot be identified without confirmed unilateral status"],
    ]
    write_xlsx(cohort / "cohort_construction_audit.xlsx", {
        "unresolved_index_ear": (["patient_id", "ear_rows", "affected_proxy_rows", "issue"], unresolved),
        "blocked_outputs": (["output", "status", "reason"], blocked),
    })

    missing_top = sorted(missing, key=lambda item: item[3], reverse=True)[:8]
    report = [
        "# Clinical data audit", "",
        f"- Unique patients (source-prefixed): {len(patients)}",
        f"- Unique baseline ears: {len({record['ear_id'] for record in baseline})}",
        f"- Total longitudinal ear records: {len(records)}",
        f"- Patients with repeat visits: {len(followup_patients)}",
        f"- Patient-level rows with uniquely encoded affected/index ear: {len(primary)}",
        f"- Index-ear unresolved: {len(unresolved)}",
        f"- Baseline affected-ear proxy rows: {affected}",
        "- Healthy-control ears: not identifiable from current coding",
        "- Disease-control ears: not identifiable from current coding",
        "- Unilateral/bilateral patient counts: not identifiable from current coding",
        "", "## Highest missingness rates",
        *[f"- {item[0]}: {item[3]:.1%} ({item[2]}/{len(records)})" for item in missing_top],
        "",
        "`affected_ear_proxy` means AAO-HNS stage is present; it is not assumed to prove unilateral or bilateral disease. The two unilateral cohort CSV files are header-only blocked outputs, not evidence that zero eligible patients exist. Names and case numbers were excluded; source numeric IDs remain in local linkage outputs and must not be treated as public-release identifiers.",
    ]
    (output / "data_audit_report.md").write_text("\n".join(report), encoding="utf-8")
    log.info(
        "patients=%d ears=%d followup=%d primary=%d unresolved=%d",
        len(patients), len(ears), len(followup_patients), len(primary), len(unresolved),
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
