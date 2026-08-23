from __future__ import annotations
import argparse
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from mdp_utils import load_config, setup_logger, write_xlsx

EXCLUDED_DIRS = {'.venv', '.git', 'node_modules', '__pycache__', '.pytest_cache'}

def classify(path: Path, out: Path) -> tuple[str,str]:
    s=str(path).lower()
    if out in path.parents: return ('intermediate' if 'intermediate' in path.parts else 'final' if 'final' in path.parts else 'log','generated output')
    if path.suffix.lower() in {'.xlsx','.xls','.csv'}: return 'raw','clinical/tabular candidate'
    if any(str(path).lower().endswith(x) for x in ['.nii.gz','.nrrd','.mha','.mhd','.dcm','.stl','.obj','.vtk']): return 'raw','image/segmentation candidate'
    if path.suffix.lower()=='.pdf': return 'reference','paper/reference'
    return 'project','code/documentation'

def main() -> int:
    ap=argparse.ArgumentParser(); ap.add_argument('--config',type=Path,required=True); a=ap.parse_args()
    _,p=load_config(a.config); log=setup_logger('inventory',p.logs/'01_inventory_files.log')
    rows=[]
    for f in p.project_root.rglob('*'):
        if not f.is_file() or EXCLUDED_DIRS.intersection(f.parts): continue
        kind,use=classify(f,p.output_root); sheets=''; dims=''
        if f.suffix.lower()=='.xlsx' and p.output_root not in f.parents:
            try:
                wb=load_workbook(f,read_only=True,data_only=False)
                sheets=';'.join(wb.sheetnames); dims=';'.join(f'{w.title}:{w.max_row}x{w.max_column}' for w in wb.worksheets)
            except Exception as e: dims=f'ERROR:{type(e).__name__}'
        st=f.stat(); rows.append([str(f.relative_to(p.project_root)),f.suffix.lower(),st.st_size,datetime.fromtimestamp(st.st_mtime).isoformat(timespec='seconds'),sheets,dims,use,kind])
    out=p.output_root/'01_data_audit'; out.mkdir(parents=True,exist_ok=True)
    headers=['relative_path','extension','bytes','modified_time','sheet_names','sheet_dimensions','possible_use','provenance_class']
    write_xlsx(out/'data_inventory.xlsx',{'inventory':(headers,rows)})
    counts={k:sum(r[-1]==k for r in rows) for k in sorted({r[-1] for r in rows})}
    report=['# Data inventory report','',f'- Files inventoried: {len(rows)}',* [f'- {k}: {v}' for k,v in counts.items()],'','Raw inputs were not modified.']
    (out/'data_audit_report.md').write_text('\n'.join(report),encoding='utf-8')
    log.info('inventoried %d files',len(rows)); return 0
if __name__=='__main__': raise SystemExit(main())
