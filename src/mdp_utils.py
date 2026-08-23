from __future__ import annotations

import csv
import hashlib
import json
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import yaml
from openpyxl import Workbook, load_workbook


REQUIRED_CONFIG = ["PROJECT_ROOT", "CLINICAL_TABLE", "SEGMENTATION_ROOT", "PEBM_PAPER", "OUTPUT_ROOT"]
ROLE_VALUES = {
    "A_static_anatomical_endotype", "B_dynamic_objective_biomarker",
    "C_fluctuating_symptom", "D_independent_validation", "E_covariate",
    "F_longitudinal_outcome", "G_identifier", "H_exclude",
}


@dataclass(frozen=True)
class ProjectPaths:
    config_path: Path
    project_root: Path
    clinical_table: Path
    segmentation_root: Path
    segmentation_batches: tuple[str, ...]
    image_root: Path | None
    pebm_paper: Path
    output_root: Path
    intermediate: Path
    final: Path
    logs: Path


def load_config(path: Path) -> tuple[dict[str, Any], ProjectPaths]:
    path = path.resolve()
    cfg = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    missing = [k for k in REQUIRED_CONFIG if not cfg.get(k)]
    if missing:
        raise ValueError(f"Missing config keys: {', '.join(missing)}")
    output = Path(cfg["OUTPUT_ROOT"])
    pp = ProjectPaths(
        path, Path(cfg["PROJECT_ROOT"]), Path(cfg["CLINICAL_TABLE"]),
        Path(cfg["SEGMENTATION_ROOT"]), tuple(cfg.get("SEGMENTATION_BATCHES", [])),
        Path(cfg["IMAGE_ROOT"]) if cfg.get("IMAGE_ROOT") else None,
        Path(cfg["PEBM_PAPER"]), output, output / "intermediate", output / "final", output / "logs",
    )
    for d in (pp.output_root, pp.intermediate, pp.final, pp.logs):
        d.mkdir(parents=True, exist_ok=True)
    return cfg, pp


def setup_logger(name: str, log_path: Path) -> logging.Logger:
    logger = logging.getLogger(name)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    for handler in (logging.FileHandler(log_path, encoding="utf-8"), logging.StreamHandler()):
        handler.setFormatter(fmt)
        logger.addHandler(handler)
    return logger


def sha256(path: Path, block: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(block), b""):
            h.update(chunk)
    return h.hexdigest()


def norm_id(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    s = str(value).strip()
    if re.fullmatch(r"\d+(\.0)?", s):
        return f"{int(float(s)):03d}"
    return s.upper().replace(" ", "")


def base_visit_id(value: Any) -> str | None:
    s = norm_id(value)
    return re.sub(r"_(?:6M|12M|\d+M|FU\d*)$", "", s or "", flags=re.I) or None


def excel_value(value: Any) -> Any:
    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, Path):
        return str(value)
    return value


def write_xlsx(path: Path, sheets: dict[str, tuple[list[str], Iterable[Iterable[Any]]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for title, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for row in rows:
            ws.append([excel_value(v) for v in row])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def write_csv(path: Path, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def read_ear_records(clinical_table: Path) -> list[dict[str, Any]]:
    wb_formula = load_workbook(clinical_table, read_only=True, data_only=False)
    wb_values = load_workbook(clinical_table, read_only=True, data_only=True)
    specs = {
        "丽水": {"site": "LS", "id": "ID", "side": "side", "freqs": ["0.5kHZ", "1kHZ", "2kHZ", "3kHZ"]},
        "浙二": {"site": "Z2", "id": "ID", "side": "side", "freqs": ["0.5kHZ", "1kHZ", "2kHZ", "3kHZ"]},
    }
    records: list[dict[str, Any]] = []
    for sheet, spec in specs.items():
        wsf, wsv = wb_formula[sheet], wb_values[sheet]
        headers = [str(c.value).strip() if c.value is not None else f"unnamed_{i}" for i, c in enumerate(next(wsf.iter_rows()), 1)]
        idx = {h: i for i, h in enumerate(headers)}
        carry: dict[str, Any] = {}
        for row_num, (rf, rv) in enumerate(zip(wsf.iter_rows(min_row=2, values_only=True), wsv.iter_rows(min_row=2, values_only=True)), 2):
            raw_id = rf[idx[spec["id"]]]
            if raw_id is not None:
                carry = {h: rf[i] for h, i in idx.items() if h not in {spec["side"], *spec["freqs"], "PTA", "CochEH", "VestEH", "VA", "ES/ED", "stage（AAO-HNS）"}}
                carry[spec["id"]] = raw_id
            visit = norm_id(carry.get(spec["id"]))
            side = str(rf[idx[spec["side"]]]).strip().upper() if rf[idx[spec["side"]]] is not None else None
            if not visit or side not in {"L", "R"}:
                continue
            patient_base = base_visit_id(visit)
            rec: dict[str, Any] = {
                "source_sheet": sheet, "source_site": spec["site"], "source_row": row_num,
                "source_subject_id": visit, "patient_id": f"{spec['site']}-{patient_base}",
                "visit_id": f"{spec['site']}-{visit}", "ear_side": side,
                "ear_id": f"{spec['site']}-{patient_base}-{side}",
            }
            for h in headers:
                if h.startswith("unnamed_") or h in {"姓名", "病案号"}:
                    continue
                val = rf[idx[h]]
                if val is None and h not in {spec["side"], *spec["freqs"], "PTA", "CochEH", "VestEH", "VA", "ES/ED", "stage（AAO-HNS）"}:
                    val = carry.get(h)
                rec[h] = val
            vals = [rv[idx[h]] for h in spec["freqs"]]
            numeric = [float(x) for x in vals if isinstance(x, (int, float))]
            rec["PTA_recomputed"] = sum(numeric) / len(numeric) if len(numeric) == 4 else None
            rec["is_followup"] = visit != patient_base
            rec["affected_ear_proxy"] = rec.get("stage（AAO-HNS）") is not None
            records.append(rec)
    return records


def provisional_variable_role(name: str) -> tuple[str, bool, bool, bool, str]:
    n = name.lower()
    if any(x in n for x in ["姓名", "电话", "病案", "出生日期", "source_row"]):
        return "H_exclude", False, False, False, "direct identifier or provenance-only field"
    if any(x in n for x in ["patient_id", "ear_id", "visit_id", "side", "id"]):
        return "G_identifier", False, False, False, "identifier/linkage field"
    if any(x in n for x in ["dhi", "thi", "耳闷", "vadl", "vertigo"]):
        return "C_fluctuating_symptom", False, True, False, "fluctuating or treatment-sensitive symptom; exclude from primary objective P-EBM"
    if any(x in n for x in ["stage", "aao-hns"]):
        return "D_independent_validation", False, False, False, "traditional clinical stage; do not reuse as both model input and primary validation outcome"
    if "es/ed" in n:
        return "B_dynamic_objective_biomarker", False, True, False, "ordinal imaging marker; coding and abnormal threshold require confirmation"
    if any(x in n for x in ["va", "semicircular", "ssc", "hsc", "psc", "morphology", "shape"]):
        return "A_static_anatomical_endotype", False, False, True, "static anatomy; not a progression event"
    if any(x in n for x in ["coch", "vest", "elh", "pta", "khz", "speech", "前庭试验"]):
        return "B_dynamic_objective_biomarker", True, True, False, "candidate only; definition/direction must be confirmed"
    if any(x in n for x in ["age", "sex", "扫描日期"]):
        return "E_covariate", False, False, False, "covariate"
    return "H_exclude", False, False, False, "requires manual classification"


def phase_stub(name: str) -> int:
    raise SystemExit(f"{name} is scaffolded for a later phase and is intentionally not run in the first round")
